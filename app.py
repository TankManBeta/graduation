from flask import Flask, render_template, request, url_for, redirect, session, jsonify, make_response, send_from_directory
from flask_login import LoginManager, login_user, login_required, current_user, logout_user
from flask_sqlalchemy import SQLAlchemy
from flask_mail import Message, Mail
from flask_caching import Cache
import settings, random, re, threading
from utils import PaperCrawler, PatentCrawler, ProjectCrawler, MyTimer, split_words
from dateutil.parser import parse
from datetime import datetime
from werkzeug.security import generate_password_hash
from urllib.parse import quote
import openpyxl, shutil, os
from openpyxl.styles import Alignment, Font
import xlrd

app = Flask(__name__)
app.config.from_object(settings)

# 数据库
db = SQLAlchemy(app)
from models import *

# 邮件
mail = Mail(app)

# 缓存
cache = Cache(app)

# 会话管理
login_manager = LoginManager()
# 绑定登陆视图的路由
login_manager.login_view = "login"
login_manager.login_message = "请您先登陆！"
login_manager.session_protection = "strong"
app.config["SECRET_KEY"] = "123456"
login_manager.init_app(app)


# 根据models里面的模型创建数据库表并生成管理员账号
# @app.before_first_request
# def create_tables():
#     db.drop_all()
#     db.create_all()
#     admin1 = User("", "管理员1", "123456")
#     db.session.add(admin1)
#     db.session.commit()
#     admin1.user_id = "admin1"
#     db.session.commit()


@login_manager.user_loader
def load_user(user_id):
    return User.query.filter_by(user_id=user_id).first()


# 处理登录的路由
@app.route('/login', methods=["GET", "POST"])
@app.route('/', methods=["GET", "POST"])
def login():
    if request.method == "GET":
        return render_template("login.html")
    if request.method == "POST":
        # 获取用户名和密码
        username = request.form.get("username")
        password = request.form.get("password")
        # 判断使用邮箱号还是账号登录
        if '@' in username:
            user = User.query.filter(User.email == username).first()
        else:
            user = User.query.filter(User.user_id == username).first()
        # 登陆成功
        if user and user.check_password(password):
            login_user(user)
            if user.user_id.startswith("admin"):
                return redirect(url_for("admin_index"))
            else:
                return redirect(url_for("index"))
        # 登录失败
        else:
            msg = "用户名或密码错误"
            return render_template("login.html", msg=msg)


# 管理员管理用户的路由
@app.route("/manage", methods=["GET", "POST"])
@login_required
def admin_index():
    if request.method == "GET":
        # 展示所有的用户信息
        users_info = db.session.query(User).filter().with_entities(User.user_id, User.user_name).all()
        users_info_after = []
        for item in users_info:
            if not item[0].startswith("admin"):
                users_info_after.append(item)
        return render_template("management.html", users_info=users_info_after)
    if request.method == "POST":
        # 获取需要重置密码的用户列表
        users_data = request.get_json()
        users_list = users_data["users_list"]
        users_password = users_data["new_password"]
        for item in users_list:
            if item == "select-all":
                continue
            else:
                user_account_info = db.session.query(User).filter(User.user_id == item).first()
                if user_account_info is not None:
                    # 重置密码，进行加密
                    user_account_info.password = generate_password_hash(users_password)
                    db.session.commit()
        return "ok"


# 登出的路由
@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


# 注册的路由
@app.route('/register', methods=["GET", "POST"])
def register():
    if request.method == "GET":
        return render_template("sign_up.html")
    if request.method == "POST":
        # 获取注册相关的信息
        email = request.form.get("email")
        password = request.form.get("password")
        password_confirmation = request.form.get("password_confirmation")
        name = request.form.get("username")
        title = request.form.get("title")
        province1 = request.form.get("province1")
        city1 = request.form.get("city1")
        district1 = request.form.get("district1")
        address1 = request.form.get("address1")
        province2 = request.form.get("province2")
        city2 = request.form.get("city2")
        district2 = request.form.get("district2")
        address2 = request.form.get("address2")
        verification = request.form.get("verification")
        # 判断邮箱是否使用过
        info = Info.query.filter(Info.email == email).first()
        if info is not None:
            error = "邮箱已被注册，请更换邮箱后重试！"
            return render_template("sign_up.html", error=error)
        else:
            # 判断两次密码是否一致
            if password != password_confirmation:
                error = "两次密码不一致，请重新注册！"
                return render_template("sign_up.html", error=error)
            else:
                # 判断通讯地址是否存在
                if province1 == "" or city1 == "" or district1 == "":
                    error = "请选择所在地区！"
                    return render_template("sign_up.html", error=error)
                elif district2 != "" and (province2 == "" or city2 == "" or district2 == ""):
                    error = "通讯地址2未选择所在地区!"
                    return render_template("sign_up.html", error=error)
                elif district2 == "" and (province2 != "" or city2 != "" or district2 != ""):
                    error = "通讯地址2未填写完整!"
                    return render_template("sign_up.html", error=error)
                else:
                    # 判断验证码是否正确
                    if cache.get(email) is None or cache.get(email) != verification:
                        error = "验证码错误，请重试！"
                        return render_template("sign_up.html", error=error)
                    else:
                        # 存储用户登录信息
                        user = User(email, name, password)
                        db.session.add(user)
                        db.session.commit()
                        # 设置用户的编号
                        user.user_id = str(user.id + 20210000)
                        db.session.commit()
                        # 存储用户信息
                        user_info = Info(name, user.user_id, email, province1, city1, district1, address1, province2,
                                         city2, district2, address2, title)
                        db.session.add(user_info)
                        db.session.commit()
                        success = "注册成功！您的账号为{}，前往".format(user.user_id)
                        # 注册成功之后系统自动开始爬取信息
                        if address2 == "":
                            address1 = split_words(address1)[0]
                            # 多线程进行爬取
                            threads = [threading.Thread(target=get_papers, args=(name, address1, user.user_id, "no")),
                                       threading.Thread(target=get_patents, args=(name, address1, user.user_id, "no")),
                                       threading.Thread(target=get_projects, args=(name, address1, user.user_id, "no"))]
                        else:
                            address1 = split_words(address1)[0]
                            address2 = split_words(address2)[0]
                            threads = [threading.Thread(target=get_papers, args=(name, address1, user.user_id, "no")),
                                       threading.Thread(target=get_patents, args=(name, address1, user.user_id, "no")),
                                       threading.Thread(target=get_projects, args=(name, address1, user.user_id, "no")),
                                       threading.Thread(target=get_papers, args=(name, address2, user.user_id, "no")),
                                       threading.Thread(target=get_patents, args=(name, address2, user.user_id, "no")),
                                       threading.Thread(target=get_projects, args=(name, address2, user.user_id, "no"))]
                        for thread in threads:
                            thread.start()
                        return render_template("sign_up.html", success=success)


# 发送验证码的路由
@app.route("/captcha", methods=["GET", "POST"])
def captcha():
    if request.method == "POST":
        # 获取邮箱
        email_data = request.get_json()
        email_account = email_data["email"]
        info = Info.query.filter(Info.email == email_account).first()
        # 判断邮箱是否已经注册过
        if info is not None:
            data = "邮箱已被注册，请更换邮箱后重试！"
        else:
            # 判断邮箱是否合法
            if re.match("^.+\\@(\\[?)[a-zA-Z0-9\\-\\.]+\\.([a-zA-Z]{2,3}|[0-9]{1,3})(\\]?)$", email_account) is None:
                data = "邮箱未填写或格式不正确"
            else:
                # 生成验证码
                code_list = []
                for i in range(6):
                    random_num = random.randint(0, 9)
                    code_list.append(str(random_num))
                verification_code = ''.join(code_list)
                # 发送邮件
                message = Message('科研信息管理系统验证码', recipients=[email_account], body='您的验证码是：%s' % verification_code)
                try:
                    # 发送
                    mail.send(message)
                    # 验证码放入缓存
                    cache.set(email_account, verification_code)
                    data = "验证码发送成功"
                except:
                    data = "验证码发送失败，请检查邮箱是否输入正确"
        return jsonify(data)


# 首页的路由
@app.route("/index", methods=["GET", "POST"])
@login_required
def index():
    if request.method == "GET":
        # 专利信息统计
        # 获取所有专利信息
        patents = db.session.query(Apply, Patent).filter\
            (Apply.teacher_id == current_user.user_id, Apply.patent_id == Patent.patent_id,
             Apply.patent_type == Patent.patent_type).with_entities(Apply.teacher_id, Apply.patent_id,
                                                                    Patent.patent_state, Patent.patent_time).all()
        # 计数
        temp_patent_count = {}
        # 构造需要返回的信息
        patent_data = [{"label": "已授权", "value": 0}, {"label": "已公开", "value": 0}, {"label": "已申请", "value": 0},
                       {"label": "已受理", "value": 0}, {"label": "已审核", "value": 0}]
        # 按状态统计
        for patent in patents:
            temp_patent_count[patent[3].year] = temp_patent_count.get(patent[3].year, 0) + 1
            if patent[2] == "已授权":
                patent_data[0]["value"] += 1
            elif patent[2] == "已公开":
                patent_data[1]["value"] += 1
            elif patent[2] == "已申请":
                patent_data[2]["value"] += 1
            elif patent[2] == "已受理":
                patent_data[3]["value"] += 1
            elif patent[2] == "已审核":
                patent_data[4]["value"] += 1
        # 论文信息统计
        # 获取所有论文信息
        papers = db.session.query(Deliver, Paper).filter\
            (Deliver.teacher_id == current_user.user_id, Paper.paper_id == Deliver.paper_id).with_entities\
            (Deliver.teacher_id, Paper.paper_id, Paper.paper_time).all()
        # 按年份计数
        temp_paper_count = {}
        for paper in papers:
            temp_paper_count[paper[2].year] = temp_paper_count.get(paper[2].year, 0) + 1
        # 构造需要返回的信息
        paper_data = []
        for key, value in temp_paper_count.items():
            res_dict = {
                "year": str(key),
                "amounts": value
            }
            paper_data.append(res_dict)
        # 项目信息统计
        projects = db.session.query(Participate, Project).filter\
            (Participate.project_id == Project.project_id, Participate.teacher_id == current_user.user_id).with_entities\
            (Participate.teacher_id, Project.project_id, Project.project_time).all()
        # 按年份计数
        temp_project_count = {}
        for project in projects:
            temp_project_count[project[2].year] = temp_project_count.get(project[2].year, 0) + 1
        # 构造返回需要的信息
        project_data = []
        for key, value in temp_project_count.items():
            project_dict = {
                "year": str(key),
                "amounts": value
            }
            project_data.append(project_dict)

        # 构造总的信息
        preview = [len(patents), len(papers), len(projects)]
        increment = [temp_patent_count.get(datetime.now().year, 0), temp_paper_count.get(datetime.now().year, 0),
                     temp_project_count.get(datetime.now().year, 0)]

        # 构造合作信息
        # 总的集合
        cooperate_dict = {}
        # 获取申请过的所有专利
        patent_id_list = db.session.query(Apply).filter(Apply.teacher_id == current_user.user_id).with_entities(
            Apply.patent_id, Apply.patent_type).all()
        # 获取这些专利的申请人
        if len(patent_id_list) != 0:
            # 存放每个专利的申请人
            patent_inventors_list = []
            for item in patent_id_list:
                inventors = db.session.query(Patent).filter(
                    Patent.patent_id == item[0], Patent.patent_type == item[1]).with_entities(Patent.patent_inventors).first()
                patent_inventors_list.append(inventors[0])
            # 做一些处理
            inventors_str = ' '.join(patent_inventors_list).replace('；', ';').replace(';', ' ')
            # 处理后的所有发明人的集合
            inventors_after = re.sub(' +', ' ', inventors_str).split(' ')
            # 开始计数
            for i in inventors_after:
                # 如果结果集合中不存在该发明人,则新建一个
                cooperate_dict[i] = cooperate_dict.get(i, {})
                cooperate_dict[i]["patents"] = cooperate_dict[i].get("patents", 0) + 1
                cooperate_dict[i]["papers"] = cooperate_dict[i].get("papers", 0)
                cooperate_dict[i]["projects"] = cooperate_dict[i].get("projects", 0)
                cooperate_dict[i]["total"] = cooperate_dict[i].get("total", 0) + 1
        # 获取所有发表过的论文
        paper_id_list = db.session.query(Deliver).filter(Deliver.teacher_id == current_user.user_id).with_entities(
            Deliver.paper_id).all()
        # 获取每篇论文的作者
        if len(paper_id_list) != 0:
            # 存放每篇论文的作者
            paper_authors_list = []
            for item in paper_id_list:
                authors = db.session.query(Paper).filter(Paper.paper_id == item[0]).with_entities(Paper.paper_authors).first()
                paper_authors_list.append(authors[0])
            # 做一些处理
            authors_str = ' '.join(paper_authors_list).replace('；', ';').replace(';', ' ')
            # 处理后的所有论文作者的集合
            authors_after = re.sub(' +', ' ', authors_str).split(' ')
            # 开始计数
            for i in authors_after:
                # 如果结果集合中不存在该作者,则新建一个
                cooperate_dict[i] = cooperate_dict.get(i, {})
                cooperate_dict[i]["patents"] = cooperate_dict[i].get("patents", 0)
                cooperate_dict[i]["papers"] = cooperate_dict[i].get("papers", 0) + 1
                cooperate_dict[i]["projects"] = cooperate_dict[i].get("projects", 0)
                cooperate_dict[i]["total"] = cooperate_dict[i].get("total", 0) + 1
        # 获取所有参与过的项目
        project_id_list = db.session.query(Participate).filter(Participate.teacher_id == current_user.user_id
                                                               ).with_entities(Participate.project_id).all()
        # 获取每个项目的主持人
        if len(project_id_list) != 0:
            # 存放各个项目的主持人
            project_principal_list = []
            for item in project_id_list:
                principal = db.session.query(Project).filter(Project.project_id == item[0]).with_entities(
                    Project.project_principal).first()
                project_principal_list.append(principal[0])
            # 做一些处理
            principals_str = ' '.join(project_principal_list).replace('；', ';').replace(';', ' ')
            # 处理后的所有项目主持人的集合
            principals_after = re.sub(' +', ' ', principals_str).split(' ')
            # 开始计数
            for i in principals_after:
                # 如果结果集合中不存在该项目负责人,则新建一个
                cooperate_dict[i] = cooperate_dict.get(i, {})
                cooperate_dict[i]["patents"] = cooperate_dict[i].get("patents", 0)
                cooperate_dict[i]["papers"] = cooperate_dict[i].get("papers", 0)
                cooperate_dict[i]["projects"] = cooperate_dict[i].get("projects", 0) + 1
                cooperate_dict[i]["total"] = cooperate_dict[i].get("total", 0) + 1
        # 筛选前五个的合作最多的
        temp_list = []
        for key in cooperate_dict.keys():
            temp_dict = {"name": key, "total": cooperate_dict[key]["total"]}
            temp_list.append(temp_dict)
        temp_list.sort(key=lambda stu: stu["total"], reverse=True)
        # 前五个的详细信息
        try:
            all_names = [temp_list[item]["name"] for item in range(1, 5)]
            all_counts = [cooperate_dict[item] for item in all_names]
            all_data = [{"name": all_names[i], "value": all_counts[i]} for i in range(0, 4)]
        except IndexError:
            all_data = []
        return render_template("index.html", patent_data=patent_data, project_data=project_data,
                               paper_data=paper_data, preview=preview, increment=increment, all_data=all_data)


# 查看详细信息的路由
@app.route("/details", methods=["GET", "POST"])
@login_required
def details():
    # 默认显示专利信息
    if request.method == "GET":
        patent_data = db.session.query(Patent, Apply).filter(
            Apply.teacher_id == current_user.user_id, Apply.patent_type == Patent.patent_type,
            Apply.patent_id == Patent.patent_id).with_entities(Patent.id, Patent.patent_id, Patent.patent_name,
                                                               Patent.patent_type, Patent.patent_owner,
                                                               Patent.patent_time, Patent.patent_state,
                                                               Patent.patent_inventors).all()
        return render_template("details_1.html", patent_data=patent_data)
    if request.method == "POST":
        # 获取筛选的条件
        data = request.get_json()
        info_type = data["info_type"]
        info_state = data["info_state"]
        start_date = data["start_date"]
        end_date = data["end_date"]
        # 全为空或全部不合法
        if start_date == "" and end_date == "":
            date_state = 0
        # 两者全都合法
        elif start_date != "" and end_date != "":
            date_state = 1
            # 按时间先后排序
            start = parse(start_date)
            end = parse(end_date)
            start_date = start if(start < end) else end
            end_date = end if (start < end) else start
        else:
            # 截止日期合法
            if start_date == "":
                date_state = 2
                end_date = parse(end_date)
            # 开始日期合法
            else:
                date_state = 3
                start_date = parse(start_date)
        # 如果要获取专利信息
        if info_type == "专利":
            # 获取全部信息
            if info_state == "全部":
                # 日期全部不合法则显示所有信息
                if date_state == 0:
                    patents = db.session.query(Patent, Apply).filter\
                        (Apply.teacher_id == current_user.user_id, Apply.patent_id == Patent.patent_id,
                         Patent.patent_type == Apply.patent_type).with_entities\
                        (Patent.id, Patent.patent_id, Patent.patent_name, Patent.patent_type, Patent.patent_owner,
                         Patent.patent_time, Patent.patent_state, Patent.patent_inventors).all()
                # 日期都合法
                elif date_state == 1:
                    patents = db.session.query(Patent, Apply).filter \
                        (Apply.teacher_id == current_user.user_id, Apply.patent_id == Patent.patent_id,
                         Patent.patent_type == Apply.patent_type, Patent.patent_time > start_date,
                         Patent.patent_time < end_date).with_entities(Patent.id, Patent.patent_id,
                                                                      Patent.patent_name, Patent.patent_type,
                                                                      Patent.patent_owner, Patent.patent_time,
                                                                      Patent.patent_state, Patent.patent_inventors).all()
                # 截止日期合法
                elif date_state == 2:
                    patents = db.session.query(Patent, Apply).filter \
                        (Apply.teacher_id == current_user.user_id, Apply.patent_id == Patent.patent_id,
                         Patent.patent_type == Apply.patent_type, Patent.patent_time < end_date).with_entities\
                        (Patent.id, Patent.patent_id, Patent.patent_name, Patent.patent_type, Patent.patent_owner,
                         Patent.patent_time, Patent.patent_state, Patent.patent_inventors).all()
                # 开始日期合法
                else:
                    patents = db.session.query(Patent, Apply).filter \
                        (Apply.teacher_id == current_user.user_id, Apply.patent_id == Patent.patent_id,
                         Patent.patent_type == Apply.patent_type, Patent.patent_time > start_date).with_entities\
                        (Patent.id, Patent.patent_id, Patent.patent_name, Patent.patent_type, Patent.patent_owner,
                         Patent.patent_time, Patent.patent_state, Patent.patent_inventors).all()
            # 获取某个状态的专利信息
            else:
                # 日期都不合法显示所有信息
                if date_state == 0:
                    patents = db.session.query(Patent, Apply).filter \
                        (Apply.teacher_id == current_user.user_id, Apply.patent_id == Patent.patent_id,
                         Patent.patent_type == Apply.patent_type, Patent.patent_state == info_state).with_entities\
                        (Patent.id, Patent.patent_id, Patent.patent_name, Patent.patent_type, Patent.patent_owner,
                         Patent.patent_time, Patent.patent_state, Patent.patent_inventors).all()
                # 日期全部合法
                elif date_state == 1:
                    patents = db.session.query(Patent, Apply).filter \
                        (Apply.teacher_id == current_user.user_id, Apply.patent_id == Patent.patent_id,
                         Patent.patent_type == Apply.patent_type, Patent.patent_state == info_state,
                         Patent.patent_time > start_date, Patent.patent_time < end_date).with_entities\
                        (Patent.id, Patent.patent_id, Patent.patent_name, Patent.patent_type, Patent.patent_owner,
                         Patent.patent_time, Patent.patent_state, Patent.patent_inventors).all()
                # 截止日期合法
                elif date_state == 2:
                    patents = db.session.query(Patent, Apply).filter \
                        (Apply.teacher_id == current_user.user_id, Apply.patent_id == Patent.patent_id,
                         Patent.patent_type == Apply.patent_type, Patent.patent_state == info_state,
                         Patent.patent_time < end_date).with_entities(Patent.id, Patent.patent_id, Patent.patent_name,
                                                                      Patent.patent_type, Patent.patent_owner,
                                                                      Patent.patent_time, Patent.patent_state,
                                                                      Patent.patent_inventors).all()
                # 开始日期合法
                else:
                    patents = db.session.query(Patent, Apply).filter \
                        (Apply.teacher_id == current_user.user_id, Apply.patent_id == Patent.patent_id,
                         Patent.patent_type == Apply.patent_type, Patent.patent_state == info_state,
                         Patent.patent_time > start_date).with_entities(Patent.id, Patent.patent_id, Patent.patent_name,
                                                                        Patent.patent_type, Patent.patent_owner,
                                                                        Patent.patent_time, Patent.patent_state,
                                                                        Patent.patent_inventors).all()
            show_html = ''
            for patent in patents:
                show_html += '<tr><td><div class="custom-control custom-checkbox"><input class="custom-control-input" '\
                             'type="checkbox" id="' + str(patent[0]) + '"><label class="custom-control-label" for="'\
                             + str(patent[0]) + '"></label></div></td><td>' + patent[1] + '</td><td>' + patent[2] + \
                             '</td><td>' + patent[3] + '</td><td>' + patent[4] + '</td><td>' + patent[-1] + '</td><td>' \
                             + patent[5].strftime('%Y-%m-%d') + '</td><td>'
                # 根据专利状态构造html
                if patent[6] == "已公开":
                    show_html += '<span class="badge badge-success">' + patent[6] + '</span></td>' + '</tr>'
                if patent[6] == "已授权":
                    show_html += '<span class="badge badge-info">' + patent[6] + '</span></td>' + '</tr>'
                if patent[6] == "已申请":
                    show_html += '<span class="badge badge-primary">' + patent[6] + '</span></td>' + '</tr>'
                if patent[6] == "已受理":
                    show_html += '<span class="badge badge-secondary">' + patent[6] + '</span></td>' + '</tr>'
                if patent[6] == "已审核":
                    show_html += '<span class="badge badge-warning">' + patent[6] + '</span></td>' + '</tr>'
            table_data = {
                "type": 0,
                "html": show_html
            }
            return jsonify(table_data)
        # 获取论文信息
        elif info_type == "论文":
            if info_state == "全部":
                # 日期全部不合法则显示所有信息
                if date_state == 0:
                    papers = db.session.query(Paper, Deliver).filter(current_user.user_id == Deliver.teacher_id,
                                                                     Deliver.paper_id == Paper.paper_id).with_entities(
                        Paper.id, Paper.paper_id, Paper.paper_name, Paper.paper_source, Paper.paper_time,
                        Paper.paper_region, Paper.paper_keywords, Paper.paper_quote, Paper.paper_influence,
                        Paper.paper_search_type, Paper.paper_press, Paper.paper_doi, Paper.paper_state,
                        Paper.paper_authors).all()
                # 日期都合法
                elif date_state == 1:
                    papers = db.session.query(Paper, Deliver).filter(current_user.user_id == Deliver.teacher_id,
                                                                     Deliver.paper_id == Paper.paper_id,
                                                                     Paper.paper_time < end_date,
                                                                     Paper.paper_time > start_date).with_entities(
                        Paper.id, Paper.paper_id, Paper.paper_name, Paper.paper_source, Paper.paper_time,
                        Paper.paper_region, Paper.paper_keywords, Paper.paper_quote, Paper.paper_influence,
                        Paper.paper_search_type, Paper.paper_press, Paper.paper_doi, Paper.paper_state,
                        Paper.paper_authors).all()
                # 截止日期合法
                elif date_state == 2:
                    papers = db.session.query(Paper, Deliver).filter(current_user.user_id == Deliver.teacher_id,
                                                                     Deliver.paper_id == Paper.paper_id,
                                                                     Paper.paper_time < end_date).with_entities(
                        Paper.id, Paper.paper_id, Paper.paper_name, Paper.paper_source, Paper.paper_time,
                        Paper.paper_region, Paper.paper_keywords, Paper.paper_quote, Paper.paper_influence,
                        Paper.paper_search_type, Paper.paper_press, Paper.paper_doi, Paper.paper_state,
                        Paper.paper_authors).all()
                # 开始日期合法
                else:
                    papers = db.session.query(Paper, Deliver).filter(current_user.user_id == Deliver.teacher_id,
                                                                     Deliver.paper_id == Paper.paper_id,
                                                                     Paper.paper_time > start_date).with_entities(
                        Paper.id, Paper.paper_id, Paper.paper_name, Paper.paper_source, Paper.paper_time,
                        Paper.paper_region, Paper.paper_keywords, Paper.paper_quote, Paper.paper_influence,
                        Paper.paper_search_type, Paper.paper_press, Paper.paper_doi, Paper.paper_state,
                        Paper.paper_authors).all()
            else:
                # 日期全部不合法则显示所有信息
                if date_state == 0:
                    papers = db.session.query(Paper, Deliver).filter(current_user.user_id == Deliver.teacher_id,
                                                                     Deliver.paper_id == Paper.paper_id,
                                                                     Paper.paper_state == info_state).with_entities(
                        Paper.id, Paper.paper_id, Paper.paper_name, Paper.paper_source, Paper.paper_time,
                        Paper.paper_region, Paper.paper_keywords, Paper.paper_quote, Paper.paper_influence,
                        Paper.paper_search_type, Paper.paper_press, Paper.paper_doi, Paper.paper_state,
                        Paper.paper_authors).all()
                # 日期都合法
                elif date_state == 1:
                    papers = db.session.query(Paper, Deliver).filter(current_user.user_id == Deliver.teacher_id,
                                                                     Deliver.paper_id == Paper.paper_id,
                                                                     Paper.paper_time < end_date,
                                                                     Paper.paper_time > start_date,
                                                                     Paper.paper_state == info_state).with_entities(
                        Paper.id, Paper.paper_id, Paper.paper_name, Paper.paper_source, Paper.paper_time,
                        Paper.paper_region, Paper.paper_keywords, Paper.paper_quote, Paper.paper_influence,
                        Paper.paper_search_type, Paper.paper_press, Paper.paper_doi, Paper.paper_state,
                        Paper.paper_authors).all()
                # 截止日期合法
                elif date_state == 2:
                    papers = db.session.query(Paper, Deliver).filter(current_user.user_id == Deliver.teacher_id,
                                                                     Deliver.paper_id == Paper.paper_id,
                                                                     Paper.paper_time < end_date,
                                                                     Paper.paper_state == info_state).with_entities(
                        Paper.id, Paper.paper_id, Paper.paper_name, Paper.paper_source, Paper.paper_time,
                        Paper.paper_region, Paper.paper_keywords, Paper.paper_quote, Paper.paper_influence,
                        Paper.paper_search_type, Paper.paper_press, Paper.paper_doi, Paper.paper_state,
                        Paper.paper_authors).all()
                # 开始日期合法
                else:
                    papers = db.session.query(Paper, Deliver).filter(current_user.user_id == Deliver.teacher_id,
                                                                     Deliver.paper_id == Paper.paper_id,
                                                                     Paper.paper_time > start_date,
                                                                     Paper.paper_state == info_state).with_entities(
                        Paper.id, Paper.paper_id, Paper.paper_name, Paper.paper_source, Paper.paper_time,
                        Paper.paper_region, Paper.paper_keywords, Paper.paper_quote, Paper.paper_influence,
                        Paper.paper_search_type, Paper.paper_press, Paper.paper_doi, Paper.paper_state,
                        Paper.paper_authors).all()
            show_html = ''
            for paper in papers:
                show_html += '<tr><td><div class="custom-control custom-checkbox"><input class="custom-control-input" '\
                             'type="checkbox" id="' + str(paper[0]) + '"><label class="custom-control-label" for="'\
                             + str(paper[0]) + '"></label></div></td><td>' + paper[1] + '</td><td>' + paper[2] + \
                             '</td><td>' + paper[3] + '</td><td>' + paper[-1] + "</td><td>" + \
                             paper[4].strftime('%Y-%m-%d') + '</td><td>' + paper[6] + '</td><td>' + str(paper[7]) \
                             + '</td><td>' + str(paper[8]) + '</td><td>' + paper[9] + '</td><td>'
                # 根据论文状态构造html
                if paper[12] == "已投递":
                    show_html += '<span class="badge badge-success">' + paper[12] + '</span></td>' + '</tr>'
                if paper[12] == "已审核":
                    show_html += '<span class="badge badge-info">' + paper[12] + '</span></td>' + '</tr>'
                if paper[12] == "已发表":
                    show_html += '<span class="badge badge-primary">' + paper[12] + '</span></td>' + '</tr>'
            table_data = {
                "type": 1,
                "html": show_html
            }
            return jsonify(table_data)
        # 获取项目信息
        else:
            # 获取全部项目信息
            if info_state == "全部":
                # 日期全部不合法则显示所有信息
                if date_state == 0:
                    projects = db.session.query(Project, Participate).filter\
                        (current_user.user_id == Participate.teacher_id, Project.project_id == Participate.project_id)\
                        .with_entities(Project.id, Project.project_id, Project.project_name, Project.project_source,
                                       Project.project_type, Project.project_principal, Project.project_principal_title,
                                       Project.project_time, Project.project_state).all()
                # 日期全部合法
                elif date_state == 1:
                    projects = db.session.query(Project, Participate).filter \
                        (current_user.user_id == Participate.teacher_id, Project.project_id == Participate.project_id,
                         Project.project_time > start_date, Project.project_time < end_date).with_entities\
                        (Project.id, Project.project_id, Project.project_name, Project.project_source,
                         Project.project_type, Project.project_principal, Project.project_principal_title,
                         Project.project_time, Project.project_state).all()
                # 截止日期合法
                elif date_state == 2:
                    projects = db.session.query(Project, Participate).filter \
                        (current_user.user_id == Participate.teacher_id, Project.project_id == Participate.project_id,
                         Project.project_time < end_date).with_entities(Project.id, Project.project_id,
                                                                        Project.project_name, Project.project_source,
                                                                        Project.project_type, Project.project_principal,
                                                                        Project.project_principal_title,
                                                                        Project.project_time, Project.project_state).all()
                # 开始日期合法
                else:
                    projects = db.session.query(Project, Participate).filter \
                        (current_user.user_id == Participate.teacher_id, Project.project_id == Participate.project_id,
                         Project.project_time > start_date).with_entities(Project.id, Project.project_id,
                                                                          Project.project_name, Project.project_source,
                                                                          Project.project_type,
                                                                          Project.project_principal,
                                                                          Project.project_principal_title,
                                                                          Project.project_time, Project.project_state).all()
            # 获取某个状态的信息
            else:
                # 日期全部不合法则显示所有信息
                if date_state == 0:
                    projects = db.session.query(Project, Participate).filter \
                        (current_user.user_id == Participate.teacher_id, Project.project_id == Participate.project_id,
                         Project.project_state == info_state).with_entities\
                        (Project.id, Project.project_id, Project.project_name, Project.project_source,
                         Project.project_type, Project.project_principal, Project.project_principal_title,
                         Project.project_time, Project.project_state).all()
                # 日期全部合法
                elif date_state == 1:
                    projects = db.session.query(Project, Participate).filter \
                        (current_user.user_id == Participate.teacher_id, Project.project_id == Participate.project_id,
                         Project.project_time > start_date, Project.project_time < end_date, Project.project_state ==
                         info_state).with_entities(Project.id, Project.project_id, Project.project_name,
                                                   Project.project_source, Project.project_type,
                                                   Project.project_principal, Project.project_principal_title,
                                                   Project.project_time, Project.project_state).all()
                # 截止日期合法
                elif date_state == 2:
                    projects = db.session.query(Project, Participate).filter \
                        (current_user.user_id == Participate.teacher_id, Project.project_id == Participate.project_id,
                         Project.project_time < end_date, Project.project_state == info_state).with_entities\
                        (Project.id, Project.project_id, Project.project_name, Project.project_source,
                         Project.project_type, Project.project_principal, Project.project_principal_title,
                         Project.project_time, Project.project_state).all()
                # 开始日期合法
                else:
                    projects = db.session.query(Project, Participate).filter \
                        (current_user.user_id == Participate.teacher_id, Project.project_id == Participate.project_id,
                         Project.project_time > start_date, Project.project_state == info_state).with_entities\
                        (Project.id, Project.project_id, Project.project_name, Project.project_source,
                         Project.project_type, Project.project_principal, Project.project_principal_title,
                         Project.project_time, Project.project_state).all()
            show_html = ''
            for project in projects:
                show_html += '<tr><td><div class="custom-control custom-checkbox"><input class="custom-control-input" ' \
                             'type="checkbox" id="' + str(project[0]) + '"><label class="custom-control-label" for="' \
                             + str(project[0]) + '"></label></div></td><td>' + project[1] + '</td><td>' + project[2] + \
                             '</td><td>' + project[3] + '</td><td>' + project[4] + '</td><td>' + project[5] + \
                             '</td><td>' + project[6] + '</td><td>' + project[7].strftime('%Y-%m-%d') + '</td><td>'
                # 根据项目状态构造html
                if project[8] == "已申请":
                    show_html += '<span class="badge badge-warning">' + project[8] + '</span></td>' + '</tr>'
                if project[8] == "已审核":
                    show_html += '<span class="badge badge-info">' + project[8] + '</span></td>' + '</tr>'
                if project[8] == "已立项":
                    show_html += '<span class="badge badge-primary">' + project[8] + '</span></td>' + '</tr>'
                if project[8] == "已结题":
                    show_html += '<span class="badge badge-success">' + project[8] + '</span></td>' + '</tr>'
            table_data = {
                "type": 2,
                "html": show_html
            }
            return jsonify(table_data)


# 删除专利的路由
@app.route("/patent/delete", methods=["GET", "POST"])
@login_required
def delete_patent():
    if request.method == "POST":
        data = request.get_json()["delete_number"]
        for item in data:
            # 全部删除
            if item == "select-all":
                continue
            # 删除部分
            else:
                patent_info = db.session.query(Patent).filter(Patent.id == item).with_entities(Patent.patent_id,
                                                                                               Patent.patent_type).first()
                if len(patent_info) != 0:
                    res = db.session.query(Apply).filter(current_user.user_id == Apply.teacher_id,
                                                         Apply.patent_id == patent_info[0],
                                                         Apply.patent_type == patent_info[1]).first()
                    db.session.delete(res)
                    db.session.commit()
        return ""


# 删除论文的路由
@app.route("/paper/delete", methods=["GET", "POST"])
def delete_paper():
    if request.method == "POST":
        data = request.get_json()["delete_number"]
        for item in data:
            # 删除所有
            if item == "select-all":
                continue
            # 删除部分
            else:
                paper_info = db.session.query(Paper).filter(Paper.id == item).with_entities(Paper.paper_id).first()
                if len(paper_info) != 0:
                    res = db.session.query(Deliver).filter(current_user.user_id == Deliver.teacher_id,
                                                           Deliver.paper_id == paper_info[0]).first()
                    db.session.delete(res)
                    db.session.commit()
        return ""


# 删除项目的路由
@app.route("/project/delete", methods=["GET", "POST"])
def delete_project():
    if request.method == "POST":
        data = request.get_json()["delete_number"]
        for item in data:
            # 删除所有
            if item == "select-all":
                continue
            # 删除部分
            else:
                project_info = db.session.query(Project).filter(Project.id == item).with_entities\
                    (Project.project_id).first()
                if len(project_info) != 0:
                    res = db.session.query(Participate).filter(current_user.user_id == Participate.teacher_id,
                                                               Participate.project_id == project_info[0]).first()
                    db.session.delete(res)
                    db.session.commit()
        return ""


# 自动更新的路由
@app.route("/update/auto", methods=["GET", "POST"])
@login_required
def update_auto():
    global timer1, timer2
    if request.method == "GET":
        return render_template("auto_update.html")
    if request.method == "POST":
        # 获取更新周期以及是否通知的选项
        data = request.get_json()
        # interval = int(data["interval"])*24*60*60
        interval = 10
        is_inform = data["is_inform"]
        # 获取用户的相关信息用于爬虫爬取
        user_data = db.session.query(Info).filter(current_user.user_id == Info.user_id).with_entities(
            Info.name, Info.address1, Info.address2, Info.user_id).first()
        name = user_data[0]
        address1 = user_data[1]
        address1 = split_words(address1)[0]
        address2 = user_data[2]
        # 通讯地址有多个
        if address2 != '':
            address2 = split_words(address2)[0]
            # 自定义两个定时器定时爬取信息
            timer1 = MyTimer(current_user.user_id, interval, update_info, (interval, name, address1, user_data[3], is_inform))
            timer2 = MyTimer(current_user.user_id, interval, update_info, (interval, name, address2, user_data[3], is_inform))
            # 启动定时任务
            timer1.start()
            timer2.start()
        else:
            timer1 = MyTimer(current_user.user_id, interval, update_info, (interval, name, address1, user_data[3], is_inform))
            # 启动定时任务
            timer1.start()
        return jsonify("success")


# 手工更新的路由
@app.route("/update/manual", methods=["GET", "POST"])
@login_required
def update_manual():
    if request.method == "GET":
        patent_data = db.session.query(Patent, Apply).filter(
            Apply.teacher_id == current_user.user_id, Apply.patent_type == Patent.patent_type,
            Apply.patent_id == Patent.patent_id).with_entities(Patent.id, Patent.patent_id, Patent.patent_name,
                                                               Patent.patent_type, Patent.patent_owner,
                                                               Patent.patent_time, Patent.patent_state,
                                                               Patent.patent_inventors).all()
        return render_template("details_3.html", patent_data=patent_data)


# 修改某个专利的路由
@app.route("/modify/patent/<patent_id>", methods=["GET", "POST"])
@login_required
def modify_patent(patent_id):
    if request.method == "GET":
        # 展示所有的专利信息
        patent_info = db.session.query(Patent).filter(Patent.id == patent_id).with_entities(
            Patent.patent_id, Patent.patent_name, Patent.patent_type, Patent.patent_state, Patent.patent_time,
            Patent.patent_owner, Patent.patent_inventors).first()
        if len(patent_info) != 0:
            teacher_apply_info = db.session.query(Apply).filter(Apply.patent_id == patent_info[0], Apply.teacher_id ==
                                                                current_user.user_id, Apply.patent_type ==
                                                                patent_info[2]).with_entities(Apply.teacher_type).first()
            # 对筛选出来的信息进行处理用于前端展示
            data = {"patent_id": patent_info[0], "patent_name": patent_info[1], "patent_type": patent_info[2],
                    "patent_state": patent_info[3], "patent_time": patent_info[4].strftime('%Y-%m-%d'),
                    "patent_owner": patent_info[5], "inventor_rank": teacher_apply_info[0],
                    'patent_inventors': patent_info[-1]}
        else:
            data = {"patent_id": "", "patent_name": "", "patent_type": "", "patent_state": "", "patent_time": "",
                    "patent_owner": "", "inventor_rank": "", "patent_inventors": ""}
        return render_template("modify_patent.html", patent_data=data)
    if request.method == "POST":
        # 获取更新的新信息
        data = request.get_json()
        print(data)
        patent_state = data["patent_state"]
        inventor_rank = data["inventor_rank"]
        patent_info = db.session.query(Patent).filter(Patent.id == patent_id).first()
        if patent_info is not None:
            # 修改信息
            patent_info.patent_state = patent_state
            db.session.commit()
            patent_type = patent_info.patent_type
            real_patent_id = patent_info.patent_id
            teacher_invent_info = db.session.query(Apply).filter(Apply.patent_id == real_patent_id, Apply.patent_type ==
                                                                 patent_type).first()
            teacher_invent_info.teacher_type = int(inventor_rank)
            # 提交更新
            db.session.commit()
            msg = "yes"
        else:
            msg = "no"
        return jsonify(msg)


# 修改某篇论文的路由
@app.route("/modify/paper/<paper_id>", methods=["GET", "POST"])
@login_required
def modify_paper(paper_id):
    if request.method == "GET":
        # 展示所有的论文信息
        paper_info = db.session.query(Paper).filter(Paper.id == paper_id).with_entities(
            Paper.paper_id, Paper.paper_name, Paper.paper_source, Paper.paper_region, Paper.paper_time,
            Paper.paper_state, Paper.paper_keywords, Paper.paper_quote, Paper.paper_influence, Paper.paper_search_type,
            Paper.paper_doi, Paper.paper_authors).first()
        if len(paper_info) != 0:
            author_rank = db.session.query(Deliver).filter(Deliver.paper_id == paper_info[0], Deliver.teacher_id ==
                                                           current_user.user_id).with_entities(Deliver.teacher_type).first()
            # 对筛选出来的信息进行处理
            data = {
                "paper_id": paper_info[0], "paper_name": paper_info[1], "paper_source": paper_info[2],
                "paper_region": paper_info[3], "paper_time": paper_info[4], "paper_state": paper_info[5],
                "paper_keywords": paper_info[6], "paper_quote": paper_info[7], "paper_influence": paper_info[8],
                "paper_search_type": paper_info[9], "paper_doi": paper_info[10], "author_rank": author_rank[0],
                "paper_authors": paper_info[-1]
            }
        else:
            # 信息为空
            data = {
                "paper_id": "", "paper_name": "", "paper_source": "", "paper_region": "", "paper_time": "",
                "paper_state": "", "paper_keywords": "", "paper_quote": "", "paper_influence": "",
                "paper_search_type": "", "paper_doi": "", "author_rank": "", "paper_authors": ""
            }
        return render_template("modify_paper.html", paper_data=data)
    if request.method == "POST":
        data = request.get_json()
        # 获取返回的信息
        paper_state = data["paper_state"]
        paper_quote = data["paper_quote"]
        paper_influence = data["paper_influence"]
        paper_search_type = data["paper_search_type"]
        paper_press = data["paper_press"]
        paper_doi = data["paper_doi"]
        author_rank = data["author_rank"]
        # 查询信息
        paper_info = db.session.query(Paper).filter(Paper.id == paper_id).first()
        # 修改信息
        if paper_info is not None:
            try:
                author_rank_info = db.session.query(Deliver).filter(Deliver.paper_id == paper_info.paper_id,
                                                                    Deliver.teacher_id == current_user.user_id).first()
                # 对信息进行更新
                paper_info.paper_state = paper_state
                paper_info.paper_quote = paper_quote
                paper_info.paper_influence = paper_influence
                paper_info.paper_search_type = paper_search_type
                paper_info.paper_press = paper_press
                paper_info.paper_doi = paper_doi
                author_rank_info.teacher_type = int(author_rank)
                # 提交信息
                db.session.commit()
                msg = "yes"
            except:
                msg = "no"
        else:
            msg = "yes"
        return jsonify(msg)


# 修改某个项目信息
@app.route("/modify/project/<project_id>", methods=["GET", "POST"])
@login_required
def modify_project(project_id):
    if request.method == "GET":
        # 筛选信息
        project_info = db.session.query(Project).filter(Project.id == project_id).with_entities(
            Project.project_id, Project.project_name, Project.project_source, Project.project_type,
            Project.project_time, Project.project_state, Project.project_principal, Project.project_principal_title
        ).first()
        if len(project_info) != 0:
            participator_rank = db.session.query(Participate).filter(
                Participate.project_id == project_info[0], Participate.teacher_id == current_user.user_id).with_entities(
                Participate.teacher_type).first()
            # 对筛选出来的信息进行加工
            data = {
                "project_id": project_info[0], "project_name": project_info[1], "project_source": project_info[2],
                "project_type": project_info[3], "project_time": project_info[4], "project_state": project_info[5],
                "project_principal": project_info[6], "project_principal_title": project_info[7],
                "participator_rank": participator_rank[0]
            }
        else:
            data = {
                "project_id": "", "project_name": "", "project_source": "", "project_type": "", "project_time": "",
                "project_state": "", "project_principal": "", "project_principal_title": "", "participator_rank": ""
            }
        return render_template("modify_project.html", project_data=data)
    if request.method == "POST":
        # 获取新的项目信息
        data = request.get_json()
        project_state = data["project_state"]
        project_principal_title = data["project_principal_title"]
        participator_rank = data["participator_rank"]
        project_info = db.session.query(Project).filter(Project.id == project_id).first()
        if project_info is not None:
            # 更新信息
            project_info.project_state = project_state
            project_info.project_principal_title = project_principal_title
            participator_info = db.session.query(Participate).filter(Participate.project_id == project_info.project_id,
                                                                     Participate.teacher_id == current_user.user_id).first()
            participator_info.teacher_type = int(participator_rank)
            # 进行提交
            db.session.commit()
            msg = "yes"
        else:
            msg = "no"
        return jsonify(msg)


# 新增专利信息的路由
@app.route("/add/patent", methods=["GET", "POST"])
@login_required
def add_patent():
    if request.method == "GET":
        return render_template("add_patent.html")
    if request.method == "POST":
        # 获取新的专利所有信息
        data = request.get_json()
        print(data)
        patent_id = data["patent_id"]
        patent_name = data["patent_name"]
        patent_type = data["patent_type"]
        patent_time = data["patent_time"]
        patent_owner = data["patent_owner"]
        patent_state = data["patent_state"]
        inventor_rank = data["inventor_rank"]
        patent_inventors = data["patent_inventors"]
        patent_inventors = patent_inventors.replace('；', ';').replace(' ', '')
        if patent_id.strip() == "" or patent_name.strip() == "" or patent_time.strip() == "" or patent_owner.strip() \
                == "" or patent_inventors.strip() == "":
            msg = "信息填写不完整"
        else:
            patent_in = db.session.query(Patent).filter(Patent.patent_id == patent_id, Patent.patent_type ==
                                                        patent_type).first()
            if patent_in is None:
                # 新建一个Patent的对象
                new_patent = Patent(patent_id, patent_name, patent_owner, parse(patent_time), patent_state, patent_type,
                                    patent_inventors)
                # 提交对象
                db.session.add(new_patent)
                db.session.commit()
                # 新建一个对应关系的对象
                new_apply = Apply(current_user.user_id, patent_id, patent_type, int(inventor_rank))
                # 提交对象
                db.session.add(new_apply)
                db.session.commit()
                msg = "yes"
            else:
                apply_info_in = db.session.query(Apply).filter(Apply.patent_id == patent_id,
                                                               Apply.teacher_id == current_user.user_id,
                                                               Apply.patent_type == patent_type).first()
                if apply_info_in is None:
                    # 修改对应关系的表
                    all_inventors = patent_in.patent_inventors
                    is_in = True if current_user.user_name in all_inventors.split(';') else False
                    if not is_in:
                        patent_in.patent_inventors = all_inventors + ";" + current_user.user_name
                        db.session.commit()
                    new_apply = Apply(current_user.user_id, patent_id, patent_type, int(inventor_rank))
                    db.session.add(new_apply)
                    db.session.commit()
                    msg = "yes"
                else:
                    msg = "该专利信息已存在"
        return jsonify(msg)


# 新增论文信息的路由
@app.route("/add/paper", methods=["GET", "POST"])
@login_required
def add_paper():
    if request.method == "GET":
        return render_template("add_paper.html")
    if request.method == "POST":
        # 获取论文的相关信息
        data = request.get_json()
        print(data)
        paper_id = data["paper_id"]
        paper_name = data["paper_name"]
        paper_source = data["paper_source"]
        paper_region = data["paper_region"]
        paper_time = data["paper_time"]
        paper_state = data["paper_state"]
        paper_keywords = data["paper_keywords"]
        paper_keywords = paper_keywords.replace("；", ";").replace(" ", "")
        paper_quote = data["paper_quote"]
        paper_influence = data["paper_influence"]
        paper_search_type = data["paper_search_type"]
        paper_press = data["paper_press"]
        paper_doi = data["paper_doi"]
        author_rank = data["author_rank"]
        paper_authors = data["paper_authors"]
        # 对论文作者进行处理
        paper_authors = paper_authors.replace("；", ";").replace(" ", "")
        if paper_id.strip() == "" or paper_name.strip() == "" or paper_source.strip() == "" or paper_time.strip() == "" \
                or paper_region.strip() == "" or paper_authors.strip() == "":
            msg = "信息填写不完整"
        else:
            try:
                paper_quote = int(paper_quote)
                paper_influence = float(paper_influence)
                paper_in = db.session.query(Paper).filter(Paper.paper_id == paper_id).first()
                # 论文信息不存在，则条件一条论文信息，一条对应关系
                if paper_in is None:
                    # 新建Paper对象
                    new_paper = Paper(paper_id, paper_name, paper_source, parse(paper_time), paper_region,
                                      paper_keywords, paper_influence, paper_quote, paper_press, paper_search_type,
                                      paper_doi, paper_state, paper_authors)
                    # 提交对象
                    db.session.add(new_paper)
                    db.session.commit()
                    # 新建Deliver对象
                    new_deliver = Deliver(current_user.user_id, paper_id, int(author_rank))
                    # 提交对象
                    db.session.add(new_deliver)
                    db.session.commit()
                    msg = "yes"
                else:
                    # 论文信息已存在，则至需要添加一条对应关系
                    deliver_info_in = db.session.query(Deliver).filter(Deliver.paper_id == paper_id,
                                                                       Deliver.teacher_id == current_user.user_id).first()
                    if deliver_info_in is None:
                        new_deliver = Deliver(current_user.user_id, paper_id, int(author_rank))
                        db.session.add(new_deliver)
                        db.session.commit()
                        msg = "yes"
                    else:
                        msg = "该论文信息已存在"
            except ValueError:
                msg = "填写信息有误"
        return jsonify(msg)


# 新增项目信息的路由
@app.route("/add/project", methods=["GET", "POST"])
@login_required
def add_project():
    if request.method == "GET":
        return render_template("add_project.html")
    if request.method == "POST":
        # 获取项目的所有信息
        data = request.get_json()
        project_id = data["project_id"]
        project_name = data["project_name"]
        project_source = data["project_source"]
        project_type = data["project_type"]
        project_state = data["project_state"]
        project_time = data["project_time"]
        project_principal = data["project_principal"]
        project_principal_title = data["project_principal_title"]
        participator_rank = data["participator_rank"]
        if project_id.strip() == "" or project_name.strip() == "" or project_time.strip() == "" or project_principal.strip() == "":
            msg = "填写信息不完整"
        else:
            project_in = db.session.query(Project).filter(Project.project_id == project_id).first()
            # 项目信息不存在，添加一条项目信息，一条对应信息
            if project_in is None:
                # 新建Project对象
                new_project = Project(project_id, project_name, project_type, project_source, project_state,
                                      project_principal, project_principal_title, parse(project_time))
                # 提交对象
                db.session.add(new_project)
                db.session.commit()
                # 新建Participate对象
                new_participate = Participate(current_user.user_id, project_id, int(participator_rank))
                # 提交对象
                db.session.add(new_participate)
                db.session.commit()
                msg = "yes"
            # 项目信息已存在，则添加一条对应关系
            else:
                participate_in = db.session.query(Participate).filter(Participate.project_id == project_id,
                                                                      Participate.teacher_id == current_user.user_id).first()
                if participate_in is None:
                    new_participate = Participate(current_user.user_id, project_id, int(participator_rank))
                    db.session.add(new_participate)
                    db.session.commit()
                    msg = "yes"
                else:
                    msg = "该项目信息已存在"
        return msg


# 修改个人信息的路由
@app.route("/modify/info", methods=["GET", "POST"])
@login_required
def modify_info():
    if request.method == "GET":
        # 获取用户的个人信息
        user_info = db.session.query(Info).filter(Info.user_id == current_user.user_id).first()
        user_id = user_info.user_id
        user_email = user_info.email
        user_name = user_info.name
        province1 = user_info.province1
        city1 = user_info.city1
        district1 = user_info.district1
        address1 = user_info.address1
        province2 = user_info.province2
        city2 = user_info.city2
        district2 = user_info.district2
        address2 = user_info.address2
        user_title = user_info.title
        # 构造字典，用于信息回填
        user_data = {
            "user_id": user_id,
            "user_email": user_email,
            "user_name": user_name,
            "province1": province1,
            "city1": city1,
            "district1": district1,
            "address1": address1,
            "province2": province2,
            "city2": city2,
            "district2": district2,
            "address2": address2,
            "user_title": user_title
        }
        return render_template("modify_user.html", user_data=user_data)
    if request.method == "POST":
        # 获取新的个人信息
        data = request.get_json()
        user_title = data["user_title"]
        province1 = data["province1"]
        city1 = data["city1"]
        district1 = data["district1"]
        address1 = data["address1"]
        province2 = data["province2"]
        city2 = data["city2"]
        district2 = data["district2"]
        address2 = data["address2"]
        if address1.strip() == "" or district1 == "" or city1 == "" or province1 == "":
            msg = "通讯地址1不完整"
        else:
            if address2.strip() == "" and district2 == "" and city2 == "" and province2 == "":
                province2 = ""
                city2 = ""
                district2 = ""
                address2 = ""
                msg = "yes"
            elif address2.strip() != "" and district2 != "" and city2 != "" and province2 != "":
                msg = "yes"
            else:
                province2 = ""
                city2 = ""
                district2 = ""
                address2 = ""
                msg = "通讯地址2不完整"
            # 修改个人信息
            user_info = db.session.query(Info).filter(Info.user_id == current_user.user_id).first()
            user_info.title = user_title
            user_info.province1 = province1
            user_info.city1 = city1
            user_info.district1 = district1
            user_info.address1 = address1
            user_info.province2 = province2
            user_info.city2 = city2
            user_info.district2 = district2
            user_info.address2 = address2
            db.session.commit()
        return jsonify(msg)


# 修改密码的路由
@app.route("/modify/password", methods=["GET", "POST"])
@login_required
def modify_password():
    if request.method == "GET":
        return render_template("modify_password.html")
    if request.method == "POST":
        # 获取新密码
        data = request.get_json()
        old_password = data["old_password"]
        new_password = data["new_password"]
        confirm_password = data["confirm_password"]
        print(data)
        old_password_info = db.session.query(User).filter(User.user_id == current_user.user_id).first()
        # 对旧密码的正确性以及新密码的长度进行校验
        if old_password_info.check_password(old_password):
            if new_password == confirm_password and len(new_password) >= 6:
                old_password_info.password = generate_password_hash(new_password)
                db.session.commit()
                msg = "yes"
            elif new_password != confirm_password:
                msg = "两次密码不一致"
            elif len(new_password) < 6:
                msg = "密码长度小于6位"
            else:
                msg = "请重新检查所填信息"
        else:
            msg = "原密码错误"
        print(msg)
        return jsonify(msg)


# 导出文件的路由
@app.route("/export", methods=["GET", "POST"])
@login_required
def export_special():
    if request.method == "GET":
        patent_data = db.session.query(Patent, Apply).filter(
            Apply.teacher_id == current_user.user_id, Apply.patent_type == Patent.patent_type,
            Apply.patent_id == Patent.patent_id).with_entities(Patent.id, Patent.patent_id, Patent.patent_name,
                                                               Patent.patent_type, Patent.patent_owner,
                                                               Patent.patent_time, Patent.patent_state,
                                                               Patent.patent_inventors).all()
        # 可选的信息头
        patent_headers = ["专利权人", "发明人", "专利号", "专利名称", "专利状态", "时间", "专利类型"]
        return render_template("special_export.html", patent_data=patent_data, patent_headers=patent_headers)
    if request.method == "POST":
        data = request.get_json()
        # 需要导出的信息项
        optional_headers = data["export_header"]
        # 需要导出的编号
        export_lists = data["export_number"]
        # 导出专利信息
        if data["export_type"] == 0:
            headers = ["专利权人", "发明人", "专利号", "专利名称", "专利状态", "时间", "专利类型"]
            # 用于存放最后的结果
            data_list = []
            # 选择了需要导出的项
            if len(export_lists) != 0:
                # 筛选所有信息
                for i in export_lists:
                    source_data = db.session.query(Patent).filter(Patent.id == i).with_entities(
                        Patent.patent_owner, Patent.patent_inventors, Patent.patent_id, Patent.patent_name,
                        Patent.patent_state, Patent.patent_time, Patent.patent_type).first()
                    # 转化时间
                    new_data = [source_data[0], source_data[1], source_data[2], source_data[3], source_data[4],
                                source_data[5].strftime('%Y-%m-%d'), source_data[6]]
                    # 导出所有项
                    if len(optional_headers) == 0:
                        data_list.append(new_data)
                    # 导出指定项
                    else:
                        final_data = []
                        for item in optional_headers:
                            # 需要哪一项就取第几个
                            final_data.append(new_data[headers.index(item)])
                        data_list.append(final_data)
                # 表头为空默认导出所有
            if len(optional_headers) == 0:
                optional_headers = ["专利权人", "发明人", "专利号", "专利名称", "专利状态", "时间", "专利类型"]
        # 导出论文信息
        elif data["export_type"] == 1:
            headers = ["论文编号", "论文名称", "论文来源", "论文作者", "发表时间", "机构地区", "关键词", "刊登信息", "检索类型", "被引量", "影响因子", "doi号", "论文状态"]
            # 用于存放最后的结果
            data_list = []
            if len(export_lists) != 0:
                for i in export_lists:
                    source_data = db.session.query(Paper).filter(Paper.id == i).with_entities(
                        Paper.paper_id, Paper.paper_name, Paper.paper_source, Paper.paper_time, Paper.paper_region,
                        Paper.paper_keywords, Paper.paper_press, Paper.paper_search_type, Paper.paper_quote,
                        Paper.paper_influence, Paper.paper_doi, Paper.paper_state, Paper.paper_authors).first()
                    # 转化时间，浮点数
                    new_data = [source_data[0], source_data[1], source_data[2], source_data[-1],
                                source_data[3].strftime("%Y-%m-%d"), source_data[4], source_data[5], source_data[6],
                                source_data[7], source_data[8], float(source_data[9]), source_data[10], source_data[11]]
                    # 导出所有项
                    if len(optional_headers) == 0:
                        data_list.append(new_data)
                    # 导出指定项
                    else:
                        final_data = []
                        for item in optional_headers:
                            # 需要哪一项就取第几个
                            final_data.append(new_data[headers.index(item)])
                        data_list.append(final_data)
            if len(optional_headers) == 0:
                optional_headers = ["论文编号", "论文名称", "论文来源", "论文作者", "发表时间", "机构地区", "关键词", "刊登信息", "检索类型", "被引量", "影响因子", "doi号", "论文状态"]
        # 导出项目信息
        else:
            headers = ["项目编号", "项目名称", "项目来源", "项目类型", "时间", "项目状态", "主持人", "主持人职称"]
            # 用于存放最后的结果
            data_list = []
            if len(export_lists) != 0:
                for i in export_lists:
                    source_data = db.session.query(Project).filter(Project.id == i).with_entities(
                        Project.project_id, Project.project_name, Project.project_source, Project.project_type,
                        Project.project_time, Project.project_state, Project.project_principal,
                        Project.project_principal_title).first()
                    # 转化时间
                    new_data = [source_data[0], source_data[1], source_data[2], source_data[3], source_data[4].strftime("%Y-%m-%d"),
                                source_data[5], source_data[6], source_data[7]]
                    # 导出所有项
                    if len(optional_headers) == 0:
                        data_list.append(new_data)
                    # 导出指定的项
                    else:
                        final_data = []
                        for item in optional_headers:
                            # 需要哪一个就取哪一个
                            final_data.append(new_data[headers.index(item)])
                        data_list.append(final_data)
            if len(optional_headers) == 0:
                optional_headers = ["项目编号", "项目名称", "项目来源", "项目类型", "时间", "项目状态", "主持人", "主持人职称"]
        # 删除文件夹里所有文件
        path = ".\\files\\"
        shutil.rmtree(path)
        os.mkdir(path)
        # 创建工作簿（默认创建一个工作表）
        new_excel = openpyxl.Workbook()
        # 选中第一个工作簿
        work_sheet = new_excel.active
        # 设置表头
        for i in range(0, len(optional_headers)):
            work_sheet.cell(row=1, column=i+1, value=optional_headers[i]).alignment = Alignment(wrapText=True,
                                                                                                horizontal='center',
                                                                                                vertical='center')
        # 填写内容
        for j in range(0, len(data_list)):
            for k in range(0, len(data_list[0])):
                work_sheet.cell(row=j+2, column=k+1, value=data_list[j][k]).alignment = Alignment(wrapText=True,
                                                                                                  horizontal='center',
                                                                                                  vertical='center')
        # 用时间戳给文件命名
        now_time = datetime.now().strftime("%Y-%m-%d-%H-%M-%S").replace('-', '')
        excel_name = current_user.user_name + '_' + now_time + '_'
        # 判断是什么信息
        if data["export_type"] == 0:
            excel_name = excel_name + "专利信息.xlsx"
        elif data["export_type"] == 1:
            excel_name = excel_name + "论文信息.xlsx"
        else:
            excel_name = excel_name + "项目信息.xlsx"
        # 保存并关闭文件
        new_excel.save(".\\files\\" + excel_name)
        new_excel.close()
        # 文件传送到前端
        res = make_response(send_from_directory(".\\files", excel_name))
        res.headers['Content-Type'] = 'text/plain;charset=UTF-8'
        res.headers['filename'] = quote(excel_name.encode("utf-8"))
        return res


# 文字识图工具的路由
@app.route("/tool", methods=["GET", "POST"])
def easy_tool():
    if request.method == "GET":
        response = make_response(send_from_directory(".", "package.exe", as_attachment=True))
        response.headers['Content-Type'] = 'text/plain;charset=UTF-8'
        return response
        # os.system(".\\package.exe")
        # return ""
    if request.method == "POST":
        response = make_response(send_from_directory(".", "package.exe", as_attachment=True))
        response.headers['Content-Type'] = 'text/plain;charset=UTF-8'
        return response
        # os.system(".\\package.exe")
        # return ""


# 下载专利信息的导入模板的路由
@app.route("/download/template/patent", methods=["GET", "POST"])
@login_required
def download_patent_template():
    if request.method == "GET":
        try:
            # 从文件夹发送文件
            response = make_response(send_from_directory(".\\import_templates", "专利信息导入模板.xlsx", as_attachment=True))
            response.headers['Content-Type'] = 'text/plain;charset=UTF-8'
            return response
        except:
            return redirect(url_for("download_patent_template"))


# 下载论文信息的导入模板
@app.route("/download/template/paper", methods=["GET", "POST"])
@login_required
def download_paper_template():
    if request.method == "GET":
        try:
            # 从文件夹发送文件
            response = make_response(send_from_directory(".\\import_templates", "论文信息导入模板.xlsx", as_attachment=True))
            response.headers['Content-Type'] = 'text/plain;charset=UTF-8'
            return response
        except:
            return redirect(url_for("download_paper_template"))


# 下载项目信息的导入模板
@app.route("/download/template/project", methods=["GET", "POST"])
@login_required
def download_project_template():
    if request.method == "GET":
        try:
            # 从文件夹发送
            response = make_response(send_from_directory(".\\import_templates", "项目信息导入模板.xlsx", as_attachment=True))
            response.headers['Content-Type'] = 'text/plain;charset=UTF-8'
            return response
        except:
            return redirect(url_for("download_project_template"))


FILE_SAVE_PATH = ".\\import_files"


# 文件批量导入专利信息的路由
@app.route("/add/patents", methods=["GET", "POST"])
@login_required
def add_patents():
    if request.method == "POST":
        # 获取文件
        file = request.files.get('file')
        if file:
            filename = file.filename
            # 判断文件类型为excel文件类型
            if filename.split('.')[-1] in ["xls", "xlsx"]:
                # 用时间戳给文件命名
                now_time = datetime.now().strftime("%Y-%m-%d-%H-%M-%S").replace('-', '')
                file_path = FILE_SAVE_PATH + "\\" + now_time + '_' + filename
                file.save(file_path)
                wb = xlrd.open_workbook(file_path)
                sheet = wb.sheets()[0]
                max_rows = sheet.nrows
                # 获取表头
                headers = sheet.row_values(0)
                # 存放需要写入数据库的数据
                data_list = []
                # 存放上传文件中出错的行
                error_list = []
                # 确定上传的是专利信息
                if "（*）专利编号" in headers:
                    for i in range(1, max_rows):
                        try:
                            row_data = sheet.row_values(i)
                            # 判断信息是否完全
                            for item in row_data:
                                try:
                                    if item.strip() == "":
                                        msg = "上传的信息不完整！"
                                        return jsonify(msg)
                                except AttributeError:
                                    continue
                            # 处理发明人
                            new_column_3 = row_data[3].replace('；', ';').replace(' ', '')
                            # 处理时间
                            new_column_5 = xlrd.xldate.xldate_as_datetime(row_data[5], 0)
                            # 处理发明人排名
                            if row_data[7] == "第一发明人":
                                new_column_7 = 0
                            elif row_data[7] == "第二发明人":
                                new_column_7 = 1
                            elif row_data[7] == "第三发明人":
                                new_column_7 = 2
                            elif row_data[7] == "第四发明人":
                                new_column_7 = 3
                            else:
                                new_column_7 = 4
                            row_data = [row_data[0], row_data[1], row_data[2], new_column_3, row_data[4], new_column_5,
                                        row_data[6], new_column_7]
                            data_list.append(row_data)
                        # 日期不匹配
                        except TypeError:
                            # 数据有错误就放入记录错误所在行数的列表中
                            error_list.append(str(i+1))
                            continue
                    # 上传的文件中有错误
                    if len(error_list) != 0:
                        msg = "第" + '、'.join(error_list) + "行数据有错误"
                    # 上传的数据没错误
                    else:
                        for item in data_list:
                            # 判断该专利信息是否已经存在
                            patent_in = db.session.query(Patent).filter(Patent.patent_id == item[0],
                                                                        Patent.patent_type == item[4]).first()
                            # 专利信息不存在，则将专利信息存入Patent表，并在Apply表存放一条对应关系
                            if patent_in is None:
                                new_patent = Patent(item[0], item[1], item[2], item[5], item[6], item[4], item[3])
                                db.session.add(new_patent)
                                db.session.commit()
                                new_apply = Apply(current_user.user_id, item[0], item[4], item[7])
                                db.session.add(new_apply)
                                db.session.commit()
                            # 专利信息已经存在,则查询Apply表中的对应关系是否存在
                            else:
                                apply_in = db.session.query(Apply).filter(Apply.teacher_id == current_user.user_id,
                                                                          Apply.patent_id == item[0],
                                                                          Apply.patent_type == item[4]).first()
                                # 若对应关系不存在则加入
                                if apply_in is None:
                                    new_apply = Apply(current_user.user_id, item[0], item[4], item[7])
                                    db.session.add(new_apply)
                                    db.session.commit()
                                # 对应关系已存在
                                else:
                                    continue
                        msg = "yes"
                else:
                    msg = "上传的文件模板不正确！"
                # 将文件删除
                os.remove(file_path)
            else:
                msg = "上传的文件格式不正确！"
        else:
            msg = "文件不存在！"
        return jsonify(msg)


# 从文件批量导入论文信息的路由
@app.route("/add/papers", methods=["GET", "POST"])
@login_required
def add_papers():
    if request.method == "POST":
        # 获取文件
        file = request.files.get('file')
        if file:
            filename = file.filename
            # 判断文件类型为excel文件类型
            if filename.split('.')[-1] in ["xls", "xlsx"]:
                # 用时间戳给文件命名
                now_time = datetime.now().strftime("%Y-%m-%d-%H-%M-%S").replace('-', '')
                file_path = FILE_SAVE_PATH + "\\" + now_time + '_' + filename
                file.save(file_path)
                wb = xlrd.open_workbook(file_path)
                sheet = wb.sheets()[0]
                max_rows = sheet.nrows
                # 获取表头
                headers = sheet.row_values(0)
                # 存放需要写入数据库的数据
                data_list = []
                # 存放上传文件中出错的行
                error_list = []
                if "（*）论文编号" in headers:
                    for i in range(1, max_rows):
                        try:
                            row_data = sheet.row_values(i)
                            # 判断必要信息是否完整
                            for j in range(0, 9):
                                try:
                                    if row_data[j].strip() == "":
                                        msg = "上传的信息不完整！"
                                        return jsonify(msg)
                                except AttributeError:
                                    continue
                            # 处理论文作者
                            new_column_3 = row_data[3].replace('；', ';').replace(' ', '')
                            # 处理时间
                            new_column_4 = xlrd.xldate.xldate_as_datetime(row_data[4], 0)
                            # 处理论文作者排名
                            if row_data[8] == "通讯作者":
                                new_column_8 = 0
                            elif row_data[8] == "第一作者":
                                new_column_8 = 1
                            elif row_data[8] == "第二作者":
                                new_column_8 = 2
                            elif row_data[8] == "第三作者":
                                new_column_8 = 3
                            else:
                                new_column_8 = 4
                            # 处理关键词
                            new_column_9 = row_data[9].replace('；', ';').replace(' ', '')
                            # 处理被引量
                            if isinstance(row_data[11], str):
                                # 如果为空，默认值设为0
                                if row_data[11].replace(' ', '') == '':
                                    new_column_11 = 0
                                # 否则就是有错误
                                else:
                                    error_list.append(str(i+1))
                                    continue
                            else:
                                new_column_11 = int(row_data[11])
                            # 处理影响因子
                            if isinstance(row_data[12], str):
                                # 如果为空，默认值设为0
                                if row_data[12].replace(' ', '') == '':
                                    new_column_12 = 0.0
                                # 否则就是有错误
                                else:
                                    error_list.append(str(i+1))
                                    continue
                            else:
                                new_column_12 = row_data[12]
                            # 需要存放进数据库的数据
                            row_data = [row_data[0], row_data[1], row_data[2], new_column_3, new_column_4, row_data[5],
                                        row_data[6], row_data[7], new_column_8, new_column_9, row_data[10],
                                        new_column_11, new_column_12, row_data[13]]
                            data_list.append(row_data)
                        # 日期错误
                        except TypeError:
                            # 数据有错误就放入记录错误所在行数的列表中
                            error_list.append(str(i+1))
                            continue
                    # 上传的文件中有错误
                    if len(error_list) != 0:
                        msg = "第" + '、'.join(error_list) + "行数据有错误！"
                    # 上传的数据没错误
                    else:
                        for item in data_list:
                            # 判断该论文信息是否已经存在
                            paper_in = db.session.query(Paper).filter(Paper.paper_id == item[0]).first()
                            # 论文信息不存在，则将论文信息存入Paper表，并在Deliver表存放一条对应关系
                            if paper_in is None:
                                new_paper = Paper(item[0], item[1], item[2], item[4], item[5], item[9], item[12],
                                                  item[11], item[10], item[6], item[13], item[7], item[3])
                                db.session.add(new_paper)
                                db.session.commit()
                                new_deliver = Deliver(current_user.user_id, item[0], item[8])
                                db.session.add(new_deliver)
                                db.session.commit()
                            # 论文信息已经存在,则查询Deliver表中的对应关系是否存在
                            else:
                                deliver_in = db.session.query(Deliver).filter(Deliver.teacher_id == current_user.user_id,
                                                                              Deliver.paper_id == item[0]).first()
                                # 若对应关系不存在则加入
                                if deliver_in is None:
                                    new_deliver = Deliver(current_user.user_id, item[0], item[8])
                                    db.session.add(new_deliver)
                                    db.session.commit()
                                # 对应关系已存在
                                else:
                                    continue
                        msg = "yes"
                else:
                    msg = "上传的文件模板不正确！"
                # 将文件删除
                os.remove(file_path)
            else:
                msg = "上传的文件格式不正确！"
        else:
            msg = "文件不存在！"
        return jsonify(msg)


# 从文件批量导入项目信息的路由
@app.route("/add/projects", methods=["GET", "POST"])
@login_required
def add_projects():
    if request.method == "POST":
        # 获取文件
        file = request.files.get('file')
        if file:
            filename = file.filename
            # 判断文件类型为excel文件类型
            if filename.split('.')[-1] in ["xls", "xlsx"]:
                # 用时间戳给文件命名
                now_time = datetime.now().strftime("%Y-%m-%d-%H-%M-%S").replace('-', '')
                file_path = FILE_SAVE_PATH + "\\" + now_time + '_' + filename
                file.save(file_path)
                wb = xlrd.open_workbook(file_path)
                sheet = wb.sheets()[0]
                max_rows = sheet.nrows
                # 获取表头
                headers = sheet.row_values(0)
                # 存放需要写入数据库的数据
                data_list = []
                # 存放上传文件中出错的行
                error_list = []
                if "（*）项目编号" in headers:
                    for i in range(1, max_rows):
                        try:
                            row_data = sheet.row_values(i)
                            # 判断必要信息是否完整
                            for j in range(0, 8):
                                try:
                                    if row_data[j].strip() == "":
                                        msg = "上传的信息不完整！"
                                        return jsonify(msg)
                                except AttributeError:
                                    continue
                            # 处理时间
                            new_column_4 = xlrd.xldate.xldate_as_datetime(row_data[4], 0)
                            # 处理参与类型
                            if row_data[6] == current_user.user_name:
                                new_column_8 = 0
                            else:
                                new_column_8 = 1
                            # 需要存放进数据库的数据
                            row_data = [row_data[0], row_data[1], row_data[2], row_data[3], new_column_4, row_data[5],
                                        row_data[6], row_data[7], new_column_8]
                            data_list.append(row_data)
                        # 日期错误
                        except TypeError:
                            # 数据有错误就放入记录错误所在行数的列表中
                            error_list.append(str(i + 1))
                            continue
                    # 上传的文件中有错误
                    if len(error_list) != 0:
                        msg = "第" + '、'.join(error_list) + "行数据有错误！"
                    # 上传的数据没错误
                    else:
                        for item in data_list:
                            # 判断该项目信息是否已经存在
                            project_in = db.session.query(Project).filter(Project.project_id == item[0]).first()
                            # 项目信息不存在，则将论文信息存入Project表，并在Participate表存放一条对应关系
                            if project_in is None:
                                new_project = Project(item[0], item[1], item[3], item[2], item[5], item[6], item[7],
                                                      item[4])
                                db.session.add(new_project)
                                db.session.commit()
                                new_participate = Participate(current_user.user_id, item[0], item[8])
                                db.session.add(new_participate)
                                db.session.commit()
                            # 项目信息已经存在,则查询Participate表中的对应关系是否存在
                            else:
                                project_in = db.session.query(Participate).filter(
                                    Participate.teacher_id == current_user.user_id,
                                    Participate.project_id == item[0]).first()
                                # 若对应关系不存在则加入
                                if project_in is None:
                                    new_participate = Participate(current_user.user_id, item[0], item[8])
                                    db.session.add(new_participate)
                                    db.session.commit()
                                # 对应关系已存在
                                else:
                                    continue
                        msg = "yes"
                else:
                    msg = "上传的文件模板不正确！"
                # 将文件删除
                os.remove(file_path)
            else:
                msg = "上传的文件格式不正确！"
        else:
            msg = "文件不存在！"
        return jsonify(msg)


# 更新信息
def update_info(interval, name, address, user_id, inform):
    global timer1
    print("我在定时执行")
    try:
        get_projects(name, address, user_id, inform)
        get_papers(name, address, user_id, inform)
        get_patents(name, address, user_id, inform)
        timer1 = MyTimer(user_id, interval, update_info, (interval, name, address, user_id, inform))
        timer1.start()
    except:
        timer1.cancel()
        timer1 = MyTimer(user_id, interval, update_info, (interval, name, address, user_id, inform))
        timer1.start()


timer1 = MyTimer("", 10, update_info)
timer2 = MyTimer("", 10, update_info)


# 获取更新论文信息
def get_papers(name, address, user_id, inform):
    print("开始获取论文信息")
    paper_crawler = PaperCrawler(name, address)
    source_data = paper_crawler.get_data()
    final_data = paper_crawler.handler_paper_items(source_data)
    new_info = "no"
    for item in final_data:
        # 获取数据
        paper_id = item["filename"]
        paper_name = item["title"]
        paper_source = item["source"]
        paper_publish_time = item["publish_time"]
        paper_publish_time = parse(paper_publish_time)
        paper_keywords = item["keywords"]
        author_rank = item["author_rank"]
        authors = item["authors"]
        # 检查论文是否已经存在
        paper_in = db.session.query(Paper).filter(Paper.paper_id == paper_id).with_entities(
            Paper.paper_id, Paper.paper_name, Paper.paper_source, Paper.paper_time, Paper.paper_region,
            Paper.paper_keywords, Paper.paper_authors).all()
        # 未存在就插入
        if len(paper_in) == 0:
            paper = Paper(paper_id, paper_name, paper_source, paper_publish_time, address, paper_keywords, 0, 0, "", "",
                          "", "已发表", authors)
            db.session.add(paper)
            db.session.commit()
            # 判断是否是新信息
            new_info = "yes"
        # 已存在就更新
        else:
            # 查找已存在的信息
            exist_data = db.session.query(Paper).filter(Paper.paper_id == paper_id).with_entities\
                (Paper.paper_id, Paper.paper_name, Paper.paper_source, Paper.paper_time, Paper.paper_region,
                 Paper.paper_keywords, Paper.paper_authors).first()
            # 存在一项不一样
            if exist_data[0] != paper_id or exist_data[1] != paper_name or exist_data[2] != paper_source or \
                    exist_data[3] != paper_publish_time or exist_data[4] != address or exist_data[5] != paper_keywords\
                    or exist_data[-1] != authors:
                in_data = db.session.query(Paper).filter(Paper.paper_id == paper_id).first()
                db.session.delete(in_data)
                db.session.commit()
                paper = Paper(paper_id, paper_name, paper_source, paper_publish_time, address, paper_keywords, 0, 0, "",
                              "", "", "已发表", authors)
                db.session.add(paper)
                db.session.commit()
                new_info = "yes"
        # 检查教师论文对应信息是否存在
        deliver_in = db.session.query(Deliver).filter(Deliver.paper_id == paper_id, Deliver.teacher_id == user_id).all()
        if len(deliver_in) == 0:
            deliver = Deliver(user_id, paper_id, author_rank)
            db.session.add(deliver)
            db.session.commit()
    # 是否邮件告知
    if inform == "yes" and new_info == "yes":
        app.app_context().push()
        account = db.session.query(User).filter(User.user_id == user_id).with_entities(User.email).first()[0]
        message = Message('科研信息管理系统', recipients=[account], body='您的论文信息有更新')
        mail.send(message)
    print("论文信息获取成功")


# 获取更新专利信息
def get_patents(name, address, user_id, inform):
    print("开始获取专利信息")
    patent_crawler = PatentCrawler(name, address)
    final_data = patent_crawler.get_data()
    new_info = "no"
    for item in final_data:
        patent_id = item["patent_id"]
        patent_name = item["patent_name"]
        patent_type = item["patent_type"]
        patent_time = item["patent_time"]
        patent_inventors = item["patent_inventors"]
        patent_time = parse(patent_time)
        inventor_rank = item["inventor_rank"]
        patent_state = '已' + patent_type[2:]
        # 检查专利是否存在
        patent_in = db.session.query(Patent).filter(Patent.patent_id == patent_id, Patent.patent_type ==
                                                    patent_type).with_entities(
            Patent.patent_id, Patent.patent_name, Patent.patent_owner, Patent.patent_time, Patent.patent_state,
            Patent.patent_type, Patent.patent_inventors).all()
        # 不存在就加入
        if len(patent_in) == 0:
            patent = Patent(patent_id, patent_name, address, patent_time, patent_state, patent_type, patent_inventors)
            db.session.add(patent)
            db.session.commit()
            new_info = "yes"
        # 存在就更新
        else:
            # 获取已存在的信息
            exist_data = db.session.query(Patent).filter(Patent.patent_id == patent_id, Patent.patent_type ==
                                                         patent_type).with_entities(Patent.patent_id, Patent.patent_name,
                                                                                    Patent.patent_owner,
                                                                                    Patent.patent_time,
                                                                                    Patent.patent_state,
                                                                                    Patent.patent_type,
                                                                                    Patent.patent_inventors).first()
            # 对比是否有不同的信息
            if exist_data[0] != patent_id or exist_data[1] != patent_name or exist_data[2] != address or \
                exist_data[3] != patent_time or exist_data[4] != patent_state or exist_data[5] != patent_type or \
                    exist_data[6] != patent_inventors:
                in_data = db.session.query(Patent).filter(Patent.patent_id == patent_id, Patent.patent_type ==
                                                          patent_type).first()
                db.session.delete(in_data)
                db.session.commit()
                patent_state = '已' + patent_type[2:]
                patent = Patent(patent_id, patent_name, address, patent_time, patent_state, patent_type, patent_inventors)
                db.session.add(patent)
                db.session.commit()
                new_info = "yes"
        # 检查教师专利对应信息是否存在
        apply_in = db.session.query(Apply).filter(Apply.patent_id == patent_id, Apply.teacher_id == user_id,
                                                  Apply.patent_type == patent_type).all()
        if len(apply_in) == 0:
            apply = Apply(user_id, patent_id, patent_type, inventor_rank)
            db.session.add(apply)
            db.session.commit()
    # 是否邮件告知
    if inform == "yes" and new_info == "yes":
        app.app_context().push()
        account = db.session.query(User).filter(User.user_id == user_id).with_entities(User.email).first()[0]
        message = Message('科研信息管理系统', recipients=[account], body='您的专利信息有更新')
        mail.send(message)
    print("专利信息获取成功")


# 获取更新项目信息
def get_projects(name, address, user_id, inform):
    print("开始获取项目信息")
    project_crawler = ProjectCrawler(name, address)
    final_data = project_crawler.get_data()
    new_info = "no"
    for item in final_data:
        project_id = item["project_id"]
        project_name = item["project_name"]
        project_type = item["project_type"]
        project_source = item["project_source"]
        project_state = item["project_state"]
        project_principal = item["project_principal"]
        project_principal_title = item["project_principal_title"]
        project_time = parse(item["project_time"])
        participant_rank = item["participant_rank"]
        # 检查项目是否已存在
        project_in = db.session.query(Project).filter(Project.project_id == project_id).with_entities(
            Project.project_id, Project.project_name, Project.project_type, Project.project_source,
            Project.project_state, Project.project_principal, Project.project_principal_title, Project.project_time
        ).all()
        # 不存在就插入
        if len(project_in) == 0:
            project = Project(project_id, project_name, project_type, project_source, project_state, project_principal,
                              project_principal_title, project_time)
            db.session.add(project)
            db.session.commit()
            new_info = "yes"
        # 存在就更新
        else:
            # 获取存在的信息
            exist_data = db.session.query(Project).filter(Project.project_id == project_id).with_entities(
                Project.project_id, Project.project_name, Project.project_type, Project.project_source,
                Project.project_state, Project.project_principal, Project.project_principal_title, Project.project_time
            ).first()
            # 对比信息是否一致
            if exist_data[0] != project_id or exist_data[1] != project_name or exist_data[2] != project_type or \
                    exist_data[3] != project_source or exist_data[4] != project_state or exist_data[5] != \
                    project_principal or exist_data[6] != project_principal_title or exist_data[7] != project_time:
                # 不一致就先删除再插入
                in_data = db.session.query(Project).filter(Project.project_id == project_id).first()
                db.session.delete(in_data)
                db.session.commit()
                project = Project(project_id, project_name, project_type, project_source, project_state, project_principal,
                                  project_principal_title, project_time)
                db.session.add(project)
                db.session.commit()
                new_info = "yes"
        # 检查教师项目对应信息是否存在
        participate_in = db.session.query(Participate).filter(Participate.project_id == project_id,
                                                              Participate.teacher_id == user_id).all()
        if len(participate_in) == 0:
            participate = Participate(user_id, project_id, participant_rank)
            db.session.add(participate)
            db.session.commit()
    # 是否邮件告知
    if inform == "yes" and new_info == "yes":
        app.app_context().push()
        account = db.session.query(User).filter(User.user_id == user_id).with_entities(User.email).first()[0]
        message = Message('科研信息管理系统', recipients=[account], body='您的项目信息有更新')
        mail.send(message)
    print("获取项目信息成功")


# 程序入口
if __name__ == '__main__':
    app.run()
